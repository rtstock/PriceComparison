# -*- coding: utf-8 -*-
"""
Created on Mon Aug 10 17:23:53 2015

@author: justin.malinchak
"""
class perform:
    def __init__(self,
                event):
                         
        self.execute(event)

        
    def parse_quoted_strings_list(s):
        import csv
        return next(csv.reader([s],
            skipinitialspace=True,
            quoting=csv.QUOTE_NONNUMERIC,
            escapechar='',
            doublequote=True,
            quotechar='"'))
 

    def execute(self,event):
                   
        import config
        #print config.sourcedir
        filenametuple = ('ComericaIPCPrice','PershingIPCPrice','WFAIPCPrice')
        datestringinfilename = '201606'
        import pandas
        from openpyxl import load_workbook
        
        import os
        lst=os.listdir(config.sourcedir)
        print lst
        #print 'in buildevent',event.src_path
        for fn in lst:
            #print fn[len(fn)-5:]
            #if fn[len(fn)-4:] == '.csv':
            if fn[len(fn)-5:] == '.xlsx':
                print fn
                book = load_workbook(config.sourcedir + '\\' + fn)
                print book.worksheets.count
                writer = pandas.ExcelWriter(config.sourcedir + '\\' + fn) 
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                for ws in writer.sheets:
                    print ws
                #for mylist in filenametuple:
                #    root = fn[:len(mylist)]
                    
                #    if root == mylist:
                #        print root
                #        if datestringinfilename in fn:
                #            print config.sourcedir + '\\' + fn
                        
                        
                        

                        #data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])
                        
                        #writer.save()
                #wb = Workbook()
                
                ## grab the active worksheet
                #ws = wb.active
                
                ## Data can be assigned directly to cells
                #ws['A1'] = 42
                
                ## Rows can also be appended
                #ws.append([1, 2, 3])
                
                ## Python types will automatically be converted
                #import datetime
                #ws['A2'] = datetime.datetime.now()
                
                ## Save the file
                #wb.save("sample.xlsx")

if __name__=='__main__':
    #o = perform(symbol='NFLX',numberofweeksahead=1)
    appname = 'buildexcelfiles'
    import datetime
    mydate = datetime.datetime.today()
    print appname,'process started',str(mydate)
    o = perform(appname)
    print appname,'process completed',datetime.datetime.today()
    #main(sys.argv[1:])
